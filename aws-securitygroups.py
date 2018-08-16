import boto3
import openpyxl

## Instantiate New Excel workbook
exl_wb = openpyxl.Workbook()
def exl_save():
    exl_wb.save('AWS.xlsx')
def print_service_header(service, region):
    print("\t\tFetching ",service, " security groups in region : ", region)
########################################################################################################################

allSGAttributes = ['RegionName', 'SecurityGroupId', 'SecurityGroupName', 'SecurityGroupDescription', 'VpcId', 'Inbound-FromPort', 'Inbound-ToPort', 'Inbound-Protocol', 'ServiceMappedTo']
numberOfAllSGAttributes = len(allSGAttributes)
allSGList = []
## Create a sheet for All security groups
exl_s0_asg = exl_wb.create_sheet(index=0, title='AllSecurityGroups')
## AllSecurityGroups - Sheet header initialization
for s_col in range(1, numberOfAllSGAttributes+1):
    exl_s0_asg.cell(row=1,column=s_col).value = allSGAttributes[s_col-1]

########################################################################################################################

elbSGAttributes = ['RegionName', 'ELBName', 'DNSName', 'AvailabilityZones', 'InstanceIDs', 'SecurityGroupIDs', 'Subnets', 'VPCId']
numberOfELBSGAttributes = len(elbSGAttributes)
elbSGList= []
## Create a sheet for ELB security groups
exl_s1_elb = exl_wb.create_sheet(index=1, title='ELBSecurityGroups')
## ELBSecurityGroups - Sheet header initialization
for s_col in range(1, numberOfELBSGAttributes+1):
    exl_s1_elb.cell(row=1,column=s_col).value = elbSGAttributes[s_col-1]

########################################################################################################################

rdsSGAttributes = ['RegionName','DBInstanceID','DBName','DBPort','DBEngine','VPCSecurityGroupID','AvailabilityZone','DBSubnetGroupName','VpcId','SubnetID']
numberOfRDSSGAttributes = len(rdsSGAttributes)
rdsSGList = []
## Create a sheet for ELB security groups
exl_s2_rds = exl_wb.create_sheet(index=2, title='RDSSecurityGroups')
## RDSSecurityGroups - Sheet header initialization
for s_col in range(1, numberOfRDSSGAttributes+1):
    exl_s2_rds.cell(row=1,column=s_col).value = rdsSGAttributes[s_col-1]

########################################################################################################################

elastiCacheSGAttributes = ['RegionName','CacheClusterId', 'CacheNodeType','CacheEngine','AvailabilityZone','CacheSubnetGroupName', 'SecurityGroupId']
numberOfElastiCacheSGAttributes = len(elastiCacheSGAttributes)
elastiCacheSGList= []
## Create a sheet for ELB security groups
exl_s3_ec = exl_wb.create_sheet(index=3, title='ElastiCacheSecurityGroups')
## RDSSecurityGroups - Sheet header initialization
for s_col in range(1, numberOfElastiCacheSGAttributes+1):
    exl_s3_ec.cell(row=1,column=s_col).value = elastiCacheSGAttributes[s_col-1]

########################################################################################################################

redshiftSGAttributes = []
numberOfRedShiftSGAttributes = len(redshiftSGAttributes)
redshiftSGList= []

########################################################################################################################
def serviceMapper(sgid):
    for i in range(0,len(elbSGList)):
        tmp = elbSGList[i][elbSGAttributes[5]].split(",")
        for j in range(0,len(tmp)):
            if str(tmp[j]) == str(sgid):
                return "ELB"
    for i in range(0, len(rdsSGList)):
        tmp = rdsSGList[i][rdsSGAttributes[5]].split(",")
        for j in range(0,len(tmp)):
            if str(tmp[j]) == str(sgid):
                return "RDS"
    for i in range(0, len(elastiCacheSGList)):
        tmp = elastiCacheSGList[i][elastiCacheSGAttributes[6]].split(",")
        for j in range(0,len(tmp)):
            if str(tmp[j]) == str(sgid):
                return "ElasticCache"
    return ""
########################################################################################################################

def processAllSecurityGroupsInRegion(region_name):
    debug_allSGfile = open("debug-allSG.txt", "a")
    print_service_header("All",region_name)
    for sg in boto3.client('ec2',region_name=region_name).describe_security_groups()['SecurityGroups']:
        allGroupsDict = {}
        allGroupsDict[allSGAttributes[0]] = region_name
        allGroupsDict[allSGAttributes[1]] = sg['GroupId']
        allGroupsDict[allSGAttributes[2]] = sg['GroupName']
        allGroupsDict[allSGAttributes[3]] = sg['Description']
        allGroupsDict[allSGAttributes[4]] = sg['VpcId']
        allGroupsDict[allSGAttributes[5]] = ""
        allGroupsDict[allSGAttributes[6]] = ""
        allGroupsDict[allSGAttributes[7]] = ""
        noofrules = len(sg['IpPermissions'])
        for ruleno in range(noofrules):
            rule = sg['IpPermissions'][ruleno]
            try:
                if rule['FromPort'] == "-1":
                    allGroupsDict[allSGAttributes[5]] = "All port"
                else:
                    allGroupsDict[allSGAttributes[5]] = str(rule['FromPort'])
            except Exception:
                pass
            try:
                if rule['ToPort'] == "-1":
                    allGroupsDict[allSGAttributes[6]] = "All port"
                else:
                    allGroupsDict[allSGAttributes[6]] = str(rule['ToPort'])
            except Exception:
                pass
            try:
                if rule['IpProtocol'] == "-1":
                    allGroupsDict[allSGAttributes[7]] = "All"
                else:
                    allGroupsDict[allSGAttributes[7]] = str(rule['IpProtocol'])
            except Exception:
                pass
        allGroupsDict[allSGAttributes[8]] = serviceMapper(allGroupsDict[allSGAttributes[1]])
        ## Append dictionary to list for later processing
        allSGList.append(allGroupsDict)
        ## Write dictionary to debug file
        debug_allSGfile.write(str(allGroupsDict))
        debug_allSGfile.write("\n")

    ## Number of all SG records found in this region
    numberOfRecords_allSG = len(allSGList)
    ## Number of records in sheet
    rowsInsheet=exl_s0_asg.max_row
    ## Writing the data
    for s_row in range(rowsInsheet+1, numberOfRecords_allSG+2):
        ## From col=1 to col=total_no_of_attributes
        for s_col in range(1, numberOfAllSGAttributes+1):
            ## fetch s_row-2 record's allSGAttributes[s_col-1] attribute
            exl_s0_asg.cell(row=s_row,column=s_col).value = allSGList[s_row-2][allSGAttributes[s_col-1]]
            # print("with data: ", allSGList[s_row-2][allSGAttributes[s_col-1]])
    exl_save()

########################################################################################################################
def processELBSecurityGroupsInRegion(region_name):
    debug_elbSGfile = open("debug-ELBSG.txt", "a")
    print_service_header("ELB",region_name)
    for elb in boto3.client('elb', region_name=region_name).describe_load_balancers()['LoadBalancerDescriptions']:
        elbGroupsDict = {}
        elbGroupsDict[elbSGAttributes[0]] = region_name
        elbGroupsDict[elbSGAttributes[1]] = elb['LoadBalancerName']
        elbGroupsDict[elbSGAttributes[2]] = elb['DNSName']
        noofaz = len(elb['AvailabilityZones'])

        elbGroupsDict[elbSGAttributes[3]] = ""
        for azno in range(noofaz):
            elbGroupsDict[elbSGAttributes[3]] = elbGroupsDict[elbSGAttributes[3]] + elb['AvailabilityZones'][azno]
            if azno != noofaz-1:
                elbGroupsDict[elbSGAttributes[3]] = elbGroupsDict[elbSGAttributes[3]] + ","

        noofinstances = len(elb['Instances'])
        elbGroupsDict[elbSGAttributes[4]] = ""
        for instanceno in range(noofinstances):
            instance = elb['Instances'][instanceno]
            elbGroupsDict[elbSGAttributes[4]] = elbGroupsDict[elbSGAttributes[4]] + instance['InstanceId']
            if instanceno != noofinstances-1:
                elbGroupsDict[elbSGAttributes[4]] = elbGroupsDict[elbSGAttributes[4]] + ","

        noofsgs = len(elb['SecurityGroups'])
        elbGroupsDict[elbSGAttributes[5]] = ""
        for sgno in range(noofsgs):
            sg = elb['SecurityGroups'][sgno]
            elbGroupsDict[elbSGAttributes[5]] = elbGroupsDict[elbSGAttributes[5]] + sg
            if sgno != noofsgs-1:
                elbGroupsDict[elbSGAttributes[5]] = elbGroupsDict[elbSGAttributes[5]] + ","

        noofsubnets = len(elb['Subnets'])
        elbGroupsDict[elbSGAttributes[6]] = ""
        for subnetno in range(noofsubnets):
            subnet = elb['Subnets'][subnetno]
            elbGroupsDict[elbSGAttributes[6]] = elbGroupsDict[elbSGAttributes[6]] + subnet
            if subnetno != noofsubnets-1:
                elbGroupsDict[elbSGAttributes[6]] = elbGroupsDict[elbSGAttributes[6]] + ","
        elbGroupsDict[elbSGAttributes[7]] = elb['VPCId']
        elbSGList.append(elbGroupsDict)
        ## Write dictionary to debug file
        debug_elbSGfile.write(str(elbGroupsDict))
        debug_elbSGfile.write("\n")

    ## Number of ELB records found in this region
    numberOfRecords_elbSG = len(elbSGList)
    ## Number of records in sheet
    rowsInsheet=exl_s1_elb.max_row
    ## Writing the data
    for s_row in range(rowsInsheet+1, numberOfRecords_elbSG+2):
        ## From col=1 to col=total_no_of_attributes
        for s_col in range(1, numberOfELBSGAttributes+1):
            ## fetch s_row-2 record's elbSGAttributes[s_col-1] attribute
            exl_s1_elb.cell(row=s_row,column=s_col).value = elbSGList[s_row-2][elbSGAttributes[s_col-1]]
            # print("with data: ", elbSGList[s_row-2][elbSGAttributes[s_col-1]])
    exl_save()

########################################################################################################################
def processRDSSecurityGroupsInRegion(region_name):
    debug_rdsSGfile = open("debug-RDSSG.txt", "a")
    print_service_header("RDS",region_name)
    for rds_db_instance in boto3.client('rds', region_name=region_name).describe_db_instances()['DBInstances']:
        rdsDbInstanceSGDict = {}
        rdsDbInstanceSGDict[rdsSGAttributes[0]] = region_name
        rdsDbInstanceSGDict[rdsSGAttributes[1]] = rds_db_instance['DBInstanceIdentifier']
        try:
            rdsDbInstanceSGDict[rdsSGAttributes[2]] = rds_db_instance['DBName']
        except Exception:
            rdsDbInstanceSGDict[rdsSGAttributes[2]] = "NULL"
        rdsDbInstanceSGDict[rdsSGAttributes[3]]=rds_db_instance['Endpoint']['Port']
        rdsDbInstanceSGDict[rdsSGAttributes[4]]=rds_db_instance['Engine']
        rdsDbInstanceSGDict[rdsSGAttributes[5]]=rds_db_instance['VpcSecurityGroups'][0]['VpcSecurityGroupId']
        rdsDbInstanceSGDict[rdsSGAttributes[6]]=rds_db_instance['AvailabilityZone']
        rdsDbInstanceSGDict[rdsSGAttributes[7]]=rds_db_instance['DBSubnetGroup']['DBSubnetGroupName']
        rdsDbInstanceSGDict[rdsSGAttributes[8]]=rds_db_instance['DBSubnetGroup']['VpcId']
        noofsubnets=len(rds_db_instance['DBSubnetGroup']['Subnets'])
        rdsDbInstanceSGDict[rdsSGAttributes[9]]=""
        for subnetno in range(noofsubnets):
            snid = rds_db_instance['DBSubnetGroup']['Subnets'][subnetno]['SubnetIdentifier']
            rdsDbInstanceSGDict[rdsSGAttributes[9]]=rdsDbInstanceSGDict[rdsSGAttributes[9]]+snid
            if subnetno != noofsubnets-1:
                rdsDbInstanceSGDict[rdsSGAttributes[9]]=rdsDbInstanceSGDict[rdsSGAttributes[9]]+","
        rdsSGList.append(rdsDbInstanceSGDict)
        ## Write dictionary to debug file
        debug_rdsSGfile.write(str(rdsDbInstanceSGDict))
        debug_rdsSGfile.write("\n")

    ## Number of RDS records found in this region
    numberOfRecords_rdsSG = len(rdsSGList)
    ## Number of records in sheet
    rowsInsheet=exl_s2_rds.max_row
    ## Writing the data
    for s_row in range(rowsInsheet+1, numberOfRecords_rdsSG+2):
        ## From col=1 to col=total_no_of_attributes
        for s_col in range(1, numberOfRDSSGAttributes+1):
            ## fetch s_row-2 record's rdsSGAttributes[s_col-1] attribute
            exl_s2_rds.cell(row=s_row,column=s_col).value = rdsSGList[s_row-2][rdsSGAttributes[s_col-1]]
            # print("with data: ", rdsSGList[s_row-2][rdsSGAttributes[s_col-1]])
    exl_save()

########################################################################################################################
def processElastiCacheSecurityGroupsInRegion(region_name):
    debug_ElastiCacheSGfile = open("debug-ElastiCacheSG.txt", "a")
    print_service_header("ElastiCache",region_name)
    for cache in boto3.client('elasticache', region_name=region_name).describe_cache_clusters()['CacheClusters']:
        cacheGroupsDict= {}
        cacheGroupsDict[elastiCacheSGAttributes[0]] = region_name
        cacheGroupsDict[elastiCacheSGAttributes[1]] = cache['CacheClusterId']
        cacheGroupsDict[elastiCacheSGAttributes[2]] = cache['CacheNodeType']
        cacheGroupsDict[elastiCacheSGAttributes[3]] = cache['Engine']
        cacheGroupsDict[elastiCacheSGAttributes[4]] = cache['PreferredAvailabilityZone']
        cacheGroupsDict[elastiCacheSGAttributes[5]] = cache['CacheSubnetGroupName']

        noofsg=len(cache['SecurityGroups'])
        cacheGroupsDict[elastiCacheSGAttributes[6]]=""
        for sgno in range(noofsg):
            sgid = cache['SecurityGroups'][sgno]['SecurityGroupId']
            cacheGroupsDict[elastiCacheSGAttributes[6]]=cacheGroupsDict[elastiCacheSGAttributes[6]]+sgid
            if sgno != noofsg-1:
                cacheGroupsDict[elastiCacheSGAttributes[6]]=cacheGroupsDict[elastiCacheSGAttributes[6]]+","
        elastiCacheSGList.append(cacheGroupsDict)
        ## Write dictionary to debug file
        debug_ElastiCacheSGfile.write(str(cacheGroupsDict))
        debug_ElastiCacheSGfile.write("\n")

    ## Number of RDS records found in this region
    numberOfRecords_elasticacheSG = len(elastiCacheSGList)
    ## Number of records in sheet
    rowsInsheet=exl_s3_ec.max_row
    ## Writing the data
    for s_row in range(rowsInsheet+1, numberOfRecords_elasticacheSG+2):
        ## From col=1 to col=total_no_of_attributes
        for s_col in range(1, numberOfElastiCacheSGAttributes+1):
            ## fetch s_row-2 record's elastiCacheSGAttributes[s_col-1] attribute
            exl_s3_ec.cell(row=s_row,column=s_col).value = elastiCacheSGList[s_row-2][elastiCacheSGAttributes[s_col-1]]
            # print("with data: ", elastiCacheSGList[s_row-2][elastiCacheSGAttributes[s_col-1]])
    exl_save()

########################################################################################################################
def processRedShiftSecurityGroupsInRegion(region_name):
    print_service_header("RedShift",region_name)
    for rs in boto3.client('redshift', region_name=region_name).describe_clusters()['Clusters']:
        redshiftSGDict= {}
        print(rs)

########################################################################################################################
def main():
    print("About Script: This script will fetch security groups related details from AWS.\n")
    ## Fetch all the region names
    region_names = [x['RegionName'] for x in boto3.client('ec2').describe_regions()['Regions']]
    print("Number of Regions found: ", len(region_names))

    for region_name in region_names:
        ## Fetching all details and security groups for Elastic Load Balancer
        processELBSecurityGroupsInRegion(region_name)

    for region_name in region_names:
        ## Fetching all details and security groups for RDS DB Instance
        processRDSSecurityGroupsInRegion(region_name)

    for region_name in region_names:
        ## Fetching all details and security groups for ElastiCache
        processElastiCacheSecurityGroupsInRegion(region_name)

    for region_name in region_names:
        ## Fetching all details and security groups for RedShift
        processRedShiftSecurityGroupsInRegion(region_name)

    for region_name in region_names:
        ## Fetching all SecurityGroups details for a region
        processAllSecurityGroupsInRegion(region_name)


########################################################################################################################
if __name__ == '__main__':
    main()
